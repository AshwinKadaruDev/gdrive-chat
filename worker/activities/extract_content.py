"""Activity: extract text content from a Google Drive file based on its MIME type."""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

import httpx
from temporalio import activity
from temporalio.exceptions import ApplicationError

logger = logging.getLogger(__name__)

DRIVE_FILES_URL = "https://www.googleapis.com/drive/v3/files"

# Google Workspace MIME types
GOOGLE_DOCS = "application/vnd.google-apps.document"
GOOGLE_SHEETS = "application/vnd.google-apps.spreadsheet"
GOOGLE_SLIDES = "application/vnd.google-apps.presentation"

# Standard file types
PDF_MIME = "application/pdf"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TEXT_MIMES = {
    "text/plain",
    "text/csv",
    "text/markdown",
    "text/html",
    "application/json",
    "application/xml",
    "text/xml",
    "text/tab-separated-values",
}

IMAGE_MIMES = {
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/webp",
    "image/bmp",
    "image/svg+xml",
}


@activity.defn
async def extract_content(file_info: dict, access_token: str) -> dict:
    """Extract text from a Google Drive file.

    Returns:
        {"text": str, "extraction_type": str, "sheets": list | None}
    """
    mime_type = file_info.get("mimeType", "")
    file_id = file_info["id"]
    file_name = file_info.get("name", "unknown")
    headers = {"Authorization": f"Bearer {access_token}"}

    activity.logger.info("Extracting content from '%s' (type=%s)", file_name, mime_type)

    try:
        return await _extract_by_mime(file_id, file_name, mime_type, headers)
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 401:
            raise ApplicationError(
                "Google Drive token expired",
                type="TOKEN_EXPIRED",
                non_retryable=True,
            ) from e
        raise


async def _extract_by_mime(
    file_id: str, file_name: str, mime_type: str, headers: dict
) -> dict:
    """Dispatch extraction based on MIME type."""
    async with httpx.AsyncClient(timeout=60.0) as client:
        # Google Docs -> export as plain text
        if mime_type == GOOGLE_DOCS:
            return await _export_google_doc(client, headers, file_id)

        # Google Sheets -> export as CSV (per sheet)
        if mime_type == GOOGLE_SHEETS:
            return await _export_google_sheets(client, headers, file_id)

        # Google Slides -> export as plain text
        if mime_type == GOOGLE_SLIDES:
            return await _export_google_slides(client, headers, file_id)

        # PDF
        if mime_type == PDF_MIME:
            return await _extract_pdf(client, headers, file_id, file_name)

        # DOCX
        if mime_type == DOCX_MIME:
            return await _extract_docx(client, headers, file_id)

        # XLSX
        if mime_type == XLSX_MIME:
            return await _extract_xlsx(client, headers, file_id)

        # Images -> placeholder
        if mime_type in IMAGE_MIMES:
            return {
                "text": f"[Image: {file_name}]",
                "extraction_type": "image_placeholder",
                "sheets": None,
            }

        # Text / CSV / Markdown / etc.
        if mime_type in TEXT_MIMES or mime_type.startswith("text/"):
            return await _download_as_text(client, headers, file_id)

        # Unknown type: try downloading as text
        activity.logger.warning("Unknown MIME type '%s' for file '%s', attempting text download", mime_type, file_name)
        try:
            return await _download_as_text(client, headers, file_id)
        except Exception:
            return {
                "text": f"[Unsupported file type: {mime_type}] {file_name}",
                "extraction_type": "unsupported",
                "sheets": None,
            }


async def _export_google_doc(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Export a Google Doc as plain text."""
    url = f"{DRIVE_FILES_URL}/{file_id}/export"
    response = await client.get(url, headers=headers, params={"mimeType": "text/plain"})
    response.raise_for_status()
    return {"text": response.text, "extraction_type": "google_doc", "sheets": None}


async def _export_google_sheets(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Export a Google Sheet.

    First fetch the spreadsheet metadata to get sheet names, then export
    each sheet as CSV individually. Returns combined text and a list of
    per-sheet text.
    """
    # Get spreadsheet metadata to find sheet names
    sheets_api_url = f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}"
    meta_response = await client.get(
        sheets_api_url, headers=headers, params={"fields": "sheets.properties"}
    )

    sheets_data: list[dict[str, Any]] = []

    if meta_response.status_code == 200:
        sheet_props = meta_response.json().get("sheets", [])
        for sheet in sheet_props:
            props = sheet.get("properties", {})
            sheet_name = props.get("title", "Sheet")
            gid = props.get("sheetId", 0)

            # Export this specific sheet as CSV
            export_url = f"{DRIVE_FILES_URL}/{file_id}/export"
            csv_response = await client.get(
                export_url,
                headers=headers,
                params={"mimeType": "text/csv", "gid": str(gid)},
            )
            if csv_response.status_code == 200:
                sheets_data.append({"name": sheet_name, "text": csv_response.text})
    else:
        # Fallback: export entire workbook as single CSV
        export_url = f"{DRIVE_FILES_URL}/{file_id}/export"
        response = await client.get(
            export_url, headers=headers, params={"mimeType": "text/csv"}
        )
        response.raise_for_status()
        sheets_data.append({"name": "Sheet1", "text": response.text})

    combined_text = "\n\n".join(
        f"=== {s['name']} ===\n{s['text']}" for s in sheets_data
    )

    return {
        "text": combined_text,
        "extraction_type": "google_sheets",
        "sheets": sheets_data,
    }


async def _export_google_slides(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Export Google Slides as plain text."""
    url = f"{DRIVE_FILES_URL}/{file_id}/export"
    response = await client.get(url, headers=headers, params={"mimeType": "text/plain"})
    response.raise_for_status()
    return {"text": response.text, "extraction_type": "google_slides", "sheets": None}


async def _extract_pdf(
    client: httpx.AsyncClient, headers: dict, file_id: str, file_name: str
) -> dict:
    """Download a PDF and extract text with pypdf, falling back to a placeholder."""
    content = await _download_file_bytes(client, headers, file_id)

    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        pages_text = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages_text.append(f"[Page {i + 1}]\n{text}")

        if pages_text:
            return {
                "text": "\n\n".join(pages_text),
                "extraction_type": "pdf_text",
                "sheets": None,
            }

        # No text extracted (likely scanned) -- return OCR placeholder
        return {
            "text": f"[Scanned PDF - OCR not available: {file_name}, {len(reader.pages)} pages]",
            "extraction_type": "pdf_ocr_placeholder",
            "sheets": None,
        }

    except Exception as e:
        logger.warning("PDF extraction failed for %s: %s", file_name, e)
        return {
            "text": f"[PDF extraction failed: {file_name}]",
            "extraction_type": "pdf_error",
            "sheets": None,
        }


async def _extract_docx(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Download a DOCX and extract text with python-docx."""
    content = await _download_file_bytes(client, headers, file_id)

    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return {
        "text": "\n\n".join(paragraphs),
        "extraction_type": "docx",
        "sheets": None,
    }


async def _extract_xlsx(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Download an XLSX and extract content with openpyxl."""
    content = await _download_file_bytes(client, headers, file_id)

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_str = ",".join(str(cell) if cell is not None else "" for cell in row)
            rows.append(row_str)
        sheet_text = "\n".join(rows)
        sheets_data.append({"name": sheet_name, "text": sheet_text})

    wb.close()

    combined_text = "\n\n".join(
        f"=== {s['name']} ===\n{s['text']}" for s in sheets_data
    )

    return {
        "text": combined_text,
        "extraction_type": "xlsx",
        "sheets": sheets_data,
    }


async def _download_as_text(client: httpx.AsyncClient, headers: dict, file_id: str) -> dict:
    """Download a file and return its content as text."""
    url = f"{DRIVE_FILES_URL}/{file_id}"
    response = await client.get(url, headers=headers, params={"alt": "media"})
    response.raise_for_status()
    return {"text": response.text, "extraction_type": "text", "sheets": None}


async def _download_file_bytes(client: httpx.AsyncClient, headers: dict, file_id: str) -> bytes:
    """Download the raw bytes of a file from Google Drive."""
    url = f"{DRIVE_FILES_URL}/{file_id}"
    response = await client.get(url, headers=headers, params={"alt": "media"})
    response.raise_for_status()
    return response.content
