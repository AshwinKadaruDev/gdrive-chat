"""
Utility functions for extracting text and metadata from common file formats.
"""

from __future__ import annotations

import io


def extract_pdf_text(content: bytes) -> str:
    """
    Extract all text from a PDF file.

    Parameters
    ----------
    content:
        Raw PDF bytes.

    Returns
    -------
    str
        Concatenated text from all pages, separated by form-feed characters.
    """
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        pages.append(text)
    return "\f".join(pages)


def extract_docx_text(content: bytes) -> str:
    """
    Extract all text from a DOCX file.

    Parameters
    ----------
    content:
        Raw DOCX bytes.

    Returns
    -------
    str
        Concatenated paragraph text separated by newlines.
    """
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs: list[str] = []
    for para in doc.paragraphs:
        paragraphs.append(para.text)
    return "\n".join(paragraphs)


def extract_xlsx_metadata(content: bytes) -> list[dict]:
    """
    Extract metadata from an XLSX file.

    Returns a list of dicts, one per sheet, with keys:
    - sheet_name: str
    - row_count: int
    - column_count: int
    - headers: list[str]

    Parameters
    ----------
    content:
        Raw XLSX bytes.

    Returns
    -------
    list[dict]
        Sheet metadata list.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets_info: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        headers: list[str] = []
        if max_row > 0:
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")

        sheets_info.append(
            {
                "sheet_name": sheet_name,
                "row_count": max_row,
                "column_count": max_col,
                "headers": headers,
            }
        )

    wb.close()
    return sheets_info


def count_tokens(text: str) -> int:
    """
    Count the number of tokens in a text string using tiktoken's cl100k_base encoding.

    Parameters
    ----------
    text:
        The text to tokenize.

    Returns
    -------
    int
        Token count.
    """
    import tiktoken

    encoding = tiktoken.get_encoding("cl100k_base")
    return len(encoding.encode(text))
