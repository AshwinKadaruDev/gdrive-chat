"""
Tool executor for the FolderAgent.

Routes tool calls to the appropriate handler, injects project_id and
access_token (never from the model), and returns a formatted result
string plus a list of citations.
"""

from __future__ import annotations

import io
import json
import logging
import re
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

_INJECTION_TAG_RE = re.compile(
    r"</?(?:SYSTEM|ADMIN|INSTRUCTION|TOOL_RESULT|OVERRIDE)[^>]*>",
    re.IGNORECASE,
)


def _sanitize_for_agent(text: str) -> str:
    """Strip tags that could be used for prompt injection."""
    return _INJECTION_TAG_RE.sub("", text)


def _require_arg(tool_args: dict, name: str) -> tuple[str | None, str | None]:
    """Return (value, None) if present, or (None, error_string) if missing."""
    val = tool_args.get(name)
    if val is None or (isinstance(val, str) and not val.strip()):
        return None, f"Error: '{name}' argument is required."
    return val, None


@dataclass
class Citation:
    """Represents a source citation from a search result."""

    chunk_id: str
    file_id: str
    file_name: str
    source_url: str | None = None
    location: str | None = None
    snippet: str = ""


async def execute_tool(
    tool_name: str,
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    """
    Execute a tool by name and return (result_string, citations_list).

    project_id and access_token are always injected server-side and never
    taken from the model's arguments.
    """
    handlers = {
        "get_folder_structure": _handle_get_folder_structure,
        "get_file_metadata": _handle_get_file_metadata,
        "read_document_pages": _handle_read_document_pages,
        "get_spreadsheet_overview": _handle_get_spreadsheet_overview,
        "read_spreadsheet_rows": _handle_read_spreadsheet_rows,
        "search_spreadsheet": _handle_search_spreadsheet,
        "get_column_stats": _handle_get_column_stats,
        "report_inability": _handle_report_inability,
        "request_clarification": _handle_request_clarification,
        # Drive-only tools
        "search_drive": _handle_search_drive,
        "get_file_content": _handle_get_file_content,
        "search_within_file_text": _handle_search_within_file_text,
    }

    handler = handlers.get(tool_name)
    if handler is None:
        logger.warning("[TOOL] Unknown tool requested: %s", tool_name)
        return f"Unknown tool: {tool_name}", []

    logger.info("[TOOL] Executing %s(args=%s, project_id=%s)", tool_name, tool_args, project_id)
    try:
        result_str, citations = await handler(
            tool_args=tool_args,
            project_id=project_id,
            access_token=access_token,
            drive_service=drive_service,
            tool_cache=tool_cache,
        )
        logger.info(
            "[TOOL] %s completed: result_length=%d, citations=%d",
            tool_name, len(result_str), len(citations),
        )
        return result_str, citations
    except Exception as exc:
        truncated_args = {k: str(v)[:200] for k, v in tool_args.items()}
        logger.warning("[TOOL] %s failed (args=%s): %s: %s", tool_name, truncated_args, type(exc).__name__, exc)
        return f"Error executing {tool_name}: {type(exc).__name__}: {exc}", []


# ------------------------------------------------------------------ #
# Tool handlers
# ------------------------------------------------------------------ #


async def _handle_get_folder_structure(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    # The project_id is used as the Google Drive folder_id
    files = await drive_service.list_folder(
        folder_id=project_id, access_token=access_token
    )

    # Cache the folder tree so later search_drive calls can skip re-crawling
    if not hasattr(drive_service, "_folder_tree_cache"):
        drive_service._folder_tree_cache = {}
    drive_service._folder_tree_cache[project_id] = files

    if not files:
        return "No files found in the folder.", []

    # Build a simple tree representation
    lines: list[str] = []
    folders: dict[str, list[dict]] = {}
    root_files: list[dict] = []

    for f in files:
        parents = f.get("parents", [])
        if parents and parents[0] != project_id:
            parent_id = parents[0]
            folders.setdefault(parent_id, []).append(f)
        else:
            root_files.append(f)

    def _format_file(f: dict, indent: int = 0) -> str:
        prefix = "  " * indent
        mime = f.get("mimeType", "")
        size = f.get("size", "")
        size_str = f" ({_format_size(int(size))})" if size else ""
        is_folder = mime == "application/vnd.google-apps.folder"
        icon = "[folder]" if is_folder else "[file]"
        return f"{prefix}{icon} {_sanitize_for_agent(f.get('name', 'unknown'))}{size_str} (id: {f.get('id', '')})"

    def _render(file_list: list[dict], indent: int = 0) -> None:
        for f in file_list:
            lines.append(_format_file(f, indent))
            if f.get("mimeType") == "application/vnd.google-apps.folder":
                children = folders.get(f["id"], [])
                _render(children, indent + 1)

    _render(root_files)

    return "\n".join(lines) if lines else "Empty folder.", []


def _format_size(size_bytes: int) -> str:
    """Format bytes into a human-readable string."""
    for unit in ("B", "KB", "MB", "GB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024  # type: ignore[assignment]
    return f"{size_bytes:.1f} TB"


async def _handle_get_file_metadata(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    metadata = await drive_service.get_file_metadata(
        file_id=file_id, access_token=access_token
    )
    return json.dumps(metadata, indent=2), []


async def _handle_read_document_pages(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    start_page, err = _require_arg(tool_args, "start_page")
    if err:
        return err, []
    end_page, err = _require_arg(tool_args, "end_page")
    if err:
        return err, []

    # First get metadata to determine file type
    metadata = await drive_service.get_file_metadata(
        file_id=file_id, access_token=access_token
    )
    mime_type = metadata.get("mimeType", "")
    file_name = _sanitize_for_agent(metadata.get("name", "unknown"))

    # Check file size before downloading
    file_size = metadata.get("size")
    if file_size is not None:
        from app.dependencies import get_settings
        max_bytes = get_settings().MAX_FILE_DOWNLOAD_BYTES
        if int(file_size) > max_bytes:
            return f"File '{file_name}' is too large ({_format_size(int(file_size))}). Max: {_format_size(max_bytes)}.", []

    if "google-apps.document" in mime_type:
        # Export Google Doc as plain text
        text = await drive_service.export_google_doc(
            file_id=file_id, access_token=access_token
        )
        # Simple page approximation: ~3000 chars per page
        chars_per_page = 3000
        start_char = (start_page - 1) * chars_per_page
        end_char = end_page * chars_per_page
        page_text = text[start_char:end_char]
        return f"Content from {file_name} (pages {start_page}-{end_page}):\n\n{page_text}", []

    # Download binary file
    content = await drive_service.download_file(
        file_id=file_id, access_token=access_token
    )

    if mime_type == "application/pdf" or file_name.lower().endswith(".pdf"):
        from app.utils.file_parsers import extract_pdf_text

        full_text = extract_pdf_text(content)
        # Split by form feed or approximate pages
        pages = full_text.split("\f") if "\f" in full_text else _split_into_pages(full_text)
        selected = pages[max(0, start_page - 1) : end_page]
        page_text = "\n\n--- Page Break ---\n\n".join(selected)
        return f"Content from {file_name} (pages {start_page}-{end_page}):\n\n{page_text}", []

    if (
        mime_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        or file_name.lower().endswith(".docx")
    ):
        from app.utils.file_parsers import extract_docx_text

        full_text = extract_docx_text(content)
        pages = _split_into_pages(full_text)
        selected = pages[max(0, start_page - 1) : end_page]
        page_text = "\n\n--- Page Break ---\n\n".join(selected)
        return f"Content from {file_name} (pages {start_page}-{end_page}):\n\n{page_text}", []

    return f"Unsupported file type for page reading: {mime_type}", []


def _split_into_pages(text: str, chars_per_page: int = 3000) -> list[str]:
    """Split text into approximate pages."""
    pages: list[str] = []
    for i in range(0, len(text), chars_per_page):
        pages.append(text[i : i + chars_per_page])
    return pages if pages else [""]


async def _download_spreadsheet_bytes(
    file_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[bytes, str, str, str | None]:
    """Download spreadsheet bytes, handling native Google Sheets via export.

    Returns (content_bytes, file_name, mime_type, web_view_link).
    """
    cache_key = f"spreadsheet:{file_id}"
    if tool_cache is not None and cache_key in tool_cache:
        return tool_cache[cache_key]

    metadata = await drive_service.get_file_metadata(
        file_id=file_id, access_token=access_token
    )
    mime_type = metadata.get("mimeType", "")
    file_name = metadata.get("name", "unknown")
    web_view_link = metadata.get("webViewLink")

    # Check file size (Google Sheets exports have no 'size' — skip gracefully)
    file_size = metadata.get("size")
    if file_size is not None:
        from app.dependencies import get_settings
        max_bytes = get_settings().MAX_SPREADSHEET_BYTES
        if int(file_size) > max_bytes:
            raise ValueError(
                f"Spreadsheet '{file_name}' is too large "
                f"({_format_size(int(file_size))}). Max: {_format_size(max_bytes)}."
            )

    if "google-apps.spreadsheet" in mime_type:
        content = await drive_service.export_google_doc(
            file_id=file_id,
            access_token=access_token,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_bytes=True,
        )
    else:
        content = await drive_service.download_file(
            file_id=file_id, access_token=access_token
        )

    result = (content, file_name, mime_type, web_view_link)
    if tool_cache is not None:
        tool_cache[cache_key] = result
    return result


async def _handle_get_spreadsheet_overview(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []

    content, file_name_raw, mime_type, _web_link = await _download_spreadsheet_bytes(
        file_id, access_token, drive_service, tool_cache=tool_cache
    )
    file_name = _sanitize_for_agent(file_name_raw)

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    overview_lines: list[str] = [f"Spreadsheet: {file_name}", ""]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        overview_lines.append(f"Sheet: {sheet_name}")
        overview_lines.append(f"  Rows: {max_row}, Columns: {max_col}")

        # Get column headers (first row)
        headers: list[str] = []
        if max_row > 0:
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            overview_lines.append(f"  Headers: {', '.join(headers)}")

        # Sample data (rows 2-4)
        if max_row > 1:
            overview_lines.append("  Sample data:")
            for row_idx in range(2, min(5, max_row + 1)):
                row_vals = []
                for cell in ws[row_idx]:
                    row_vals.append(str(cell.value) if cell.value is not None else "")
                overview_lines.append(f"    Row {row_idx}: {', '.join(row_vals)}")

        overview_lines.append("")

    wb.close()
    return "\n".join(overview_lines), []


async def _handle_read_spreadsheet_rows(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    sheet_name, err = _require_arg(tool_args, "sheet_name")
    if err:
        return err, []
    start_row, err = _require_arg(tool_args, "start_row")
    if err:
        return err, []
    end_row, err = _require_arg(tool_args, "end_row")
    if err:
        return err, []

    content, file_name_raw, _mime_type, web_view_link = await _download_spreadsheet_bytes(
        file_id, access_token, drive_service, tool_cache=tool_cache
    )
    file_name = _sanitize_for_agent(file_name_raw)

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}", []

    ws = wb[sheet_name]

    # Get headers from row 1
    headers: list[str] = []
    if ws.max_row and ws.max_row > 0:
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value is not None else "")

    lines: list[str] = []
    if headers:
        lines.append(" | ".join(headers))
        lines.append("-" * len(lines[0]))

    for row_idx in range(max(1, start_row), min(end_row + 1, (ws.max_row or 0) + 1)):
        row_vals: list[str] = []
        for cell in ws[row_idx]:
            row_vals.append(str(cell.value) if cell.value is not None else "")
        lines.append(f"Row {row_idx}: {' | '.join(row_vals)}")

    wb.close()

    result_text = "\n".join(lines) if lines else "No data found in the specified range."

    # Build a snippet from the first few data lines (skip header separator)
    snippet_lines = [l for l in lines[:5] if not l.startswith("---")]
    citation = Citation(
        chunk_id=f"spreadsheet-rows-{file_id}-{sheet_name}-{start_row}-{end_row}",
        file_id=file_id,
        file_name=file_name,
        source_url=web_view_link,
        location=f"Sheet: {sheet_name}, Rows {start_row}-{end_row}",
        snippet="\n".join(snippet_lines)[:300],
    )
    return result_text, [citation]


async def _handle_search_spreadsheet(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    query, err = _require_arg(tool_args, "query")
    if err:
        return err, []
    query = query.lower()
    target_sheet = tool_args.get("sheet_name")

    content, file_name_raw, _mime_type, web_view_link = await _download_spreadsheet_bytes(
        file_id, access_token, drive_service, tool_cache=tool_cache
    )
    file_name = _sanitize_for_agent(file_name_raw)

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    matches: list[str] = []
    sheets_to_search = [target_sheet] if target_sheet else wb.sheetnames

    for sn in sheets_to_search:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]

        # Get headers
        headers: list[str] = []
        if ws.max_row and ws.max_row > 0:
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")

        from app.dependencies import get_settings
        max_rows = get_settings().MAX_SPREADSHEET_SEARCH_ROWS

        rows_scanned = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            rows_scanned += 1
            if rows_scanned > max_rows:
                matches.append(f"⚠ Stopped after scanning {max_rows} rows.")
                break
            row_vals: list[str] = []
            found = False
            for cell in ws[row_idx]:
                val = str(cell.value) if cell.value is not None else ""
                row_vals.append(val)
                if query in val.lower():
                    found = True
            if found:
                safe_vals = [_sanitize_for_agent(v) for v in row_vals]
                matches.append(f"Sheet '{sn}', Row {row_idx}: {' | '.join(safe_vals)}")

        if len(matches) >= 50:
            break

    wb.close()

    if not matches:
        return f"No matches found for '{tool_args['query']}'.", []

    result_text = f"Found {len(matches)} matching rows:\n" + "\n".join(matches)

    sheet_label = target_sheet or "all sheets"
    citation = Citation(
        chunk_id=f"spreadsheet-search-{file_id}-{tool_args['query'][:30]}",
        file_id=file_id,
        file_name=file_name,
        source_url=web_view_link,
        location=f"Search in {sheet_label}: '{tool_args['query']}'",
        snippet="\n".join(matches[:3])[:300],
    )
    return result_text, [citation]


async def _handle_get_column_stats(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    sheet_name, err = _require_arg(tool_args, "sheet_name")
    if err:
        return err, []
    column_name, err = _require_arg(tool_args, "column_name")
    if err:
        return err, []

    content, file_name_raw, _mime_type, web_view_link = await _download_spreadsheet_bytes(
        file_id, access_token, drive_service, tool_cache=tool_cache
    )
    file_name = _sanitize_for_agent(file_name_raw)

    import openpyxl
    import statistics

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"Sheet '{sheet_name}' not found.", []

    ws = wb[sheet_name]

    # Find the column index
    col_idx: int | None = None
    if ws.max_row and ws.max_row > 0:
        for i, cell in enumerate(ws[1]):
            if str(cell.value).strip().lower() == column_name.strip().lower():
                col_idx = i
                break

    if col_idx is None:
        headers = []
        if ws.max_row and ws.max_row > 0:
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
        wb.close()
        return (
            f"Column '{column_name}' not found. Available columns: {', '.join(headers)}",
            [],
        )

    # Collect numeric values
    values: list[float] = []
    for row_idx in range(2, (ws.max_row or 0) + 1):
        row_cells = list(ws[row_idx])
        if col_idx < len(row_cells):
            val = row_cells[col_idx].value
            if val is not None:
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    pass

    wb.close()

    if not values:
        return f"No numeric values found in column '{column_name}'.", []

    stats_lines = [
        f"Statistics for column '{column_name}' in sheet '{sheet_name}':",
        f"  Count: {len(values)}",
        f"  Sum: {sum(values):.2f}",
        f"  Mean: {statistics.mean(values):.2f}",
        f"  Median: {statistics.median(values):.2f}",
        f"  Min: {min(values):.2f}",
        f"  Max: {max(values):.2f}",
    ]
    if len(values) >= 2:
        stats_lines.append(f"  Std Dev: {statistics.stdev(values):.2f}")

    result_text = "\n".join(stats_lines)

    citation = Citation(
        chunk_id=f"spreadsheet-stats-{file_id}-{sheet_name}-{column_name}",
        file_id=file_id,
        file_name=file_name,
        source_url=web_view_link,
        location=f"Sheet: {sheet_name}, Column: {column_name}",
        snippet=result_text[:300],
    )
    return result_text, [citation]


async def _handle_report_inability(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    reason, err = _require_arg(tool_args, "reason")
    if err:
        return err, []
    return reason, []


async def _handle_request_clarification(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    question, err = _require_arg(tool_args, "question")
    if err:
        return err, []
    return question, []


# ------------------------------------------------------------------ #
# Drive-only tool handlers
# ------------------------------------------------------------------ #


async def _handle_search_drive(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    query, err = _require_arg(tool_args, "query")
    if err:
        return err, []
    file_types = tool_args.get("file_types")

    results = await drive_service.search_files(
        folder_id=project_id,
        query=query,
        access_token=access_token,
        file_types=file_types,
    )

    if not results:
        return "No files found matching your query.", []

    formatted_parts: list[str] = []
    for i, f in enumerate(results, 1):
        file_id = f.get("id", "")
        file_name = _sanitize_for_agent(f.get("name", "unknown"))
        mime_type = f.get("mimeType", "")
        size = f.get("size", "")
        size_str = f" ({_format_size(int(size))})" if size else ""
        modified = f.get("modifiedTime", "")

        formatted_parts.append(
            f"[Result {i}] {file_name}{size_str}\n"
            f"  ID: {file_id}\n"
            f"  Type: {mime_type}\n"
            f"  Modified: {modified}\n"
        )

    return f"Found {len(results)} files:\n\n" + "\n".join(formatted_parts), []


async def _handle_get_file_content(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    max_chars = tool_args.get("max_chars", 50000)

    metadata = await drive_service.get_file_metadata(
        file_id=file_id, access_token=access_token
    )
    mime_type = metadata.get("mimeType", "")
    file_name = _sanitize_for_agent(metadata.get("name", "unknown"))
    web_link = metadata.get("webViewLink")

    # Check file size before downloading
    file_size = metadata.get("size")
    if file_size is not None:
        from app.dependencies import get_settings
        max_bytes = get_settings().MAX_FILE_DOWNLOAD_BYTES
        if int(file_size) > max_bytes:
            return f"File '{file_name}' is too large ({_format_size(int(file_size))}). Max: {_format_size(max_bytes)}.", []

    # Google Docs/Sheets/Slides — export as plain text
    if "google-apps.document" in mime_type:
        text = await drive_service.export_google_doc(
            file_id=file_id, access_token=access_token
        )
    elif "google-apps.spreadsheet" in mime_type:
        text = await drive_service.export_google_doc(
            file_id=file_id,
            access_token=access_token,
            mime_type="text/csv",
        )
    elif "google-apps.presentation" in mime_type:
        text = await drive_service.export_google_doc(
            file_id=file_id, access_token=access_token
        )
    else:
        # Binary file — download and parse
        content = await drive_service.download_file(
            file_id=file_id, access_token=access_token
        )

        if mime_type == "application/pdf" or file_name.lower().endswith(".pdf"):
            from app.utils.file_parsers import extract_pdf_text
            text = extract_pdf_text(content)
        elif (
            mime_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or file_name.lower().endswith(".docx")
        ):
            from app.utils.file_parsers import extract_docx_text
            text = extract_docx_text(content)
        elif mime_type.startswith("text/") or file_name.lower().endswith(".txt"):
            text = content.decode("utf-8", errors="replace")
        else:
            text = f"[Unsupported file type: {mime_type}. Cannot extract text.]"

    truncated = len(text) > max_chars
    text = text[:max_chars]

    citation = Citation(
        chunk_id=f"file-content-{file_id}",
        file_id=file_id,
        file_name=file_name,
        source_url=web_link,
        location=None,
        snippet=text[:300],
    )

    header = f"Content of {file_name}"
    if truncated:
        header += f" (truncated to {max_chars} characters)"
    header += ":\n\n"

    return header + text, [citation]


async def _handle_search_within_file_text(
    tool_args: dict,
    project_id: str,
    access_token: str,
    drive_service: "GoogleDriveService",
    tool_cache: dict | None = None,
) -> tuple[str, list[Citation]]:
    file_id, err = _require_arg(tool_args, "file_id")
    if err:
        return err, []
    query, err = _require_arg(tool_args, "query")
    if err:
        return err, []
    context_chars = tool_args.get("context_chars", 200)

    metadata = await drive_service.get_file_metadata(
        file_id=file_id, access_token=access_token
    )
    mime_type = metadata.get("mimeType", "")
    file_name = _sanitize_for_agent(metadata.get("name", "unknown"))
    web_link = metadata.get("webViewLink")

    # Check file size before downloading
    file_size = metadata.get("size")
    if file_size is not None:
        from app.dependencies import get_settings
        max_bytes = get_settings().MAX_FILE_DOWNLOAD_BYTES
        if int(file_size) > max_bytes:
            return f"File '{file_name}' is too large ({_format_size(int(file_size))}). Max: {_format_size(max_bytes)}.", []

    # Get the full text
    if "google-apps.document" in mime_type:
        text = await drive_service.export_google_doc(
            file_id=file_id, access_token=access_token
        )
    elif "google-apps.spreadsheet" in mime_type:
        text = await drive_service.export_google_doc(
            file_id=file_id,
            access_token=access_token,
            mime_type="text/csv",
        )
    else:
        content = await drive_service.download_file(
            file_id=file_id, access_token=access_token
        )
        if mime_type == "application/pdf" or file_name.lower().endswith(".pdf"):
            from app.utils.file_parsers import extract_pdf_text
            text = extract_pdf_text(content)
        elif (
            mime_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or file_name.lower().endswith(".docx")
        ):
            from app.utils.file_parsers import extract_docx_text
            text = extract_docx_text(content)
        else:
            text = content.decode("utf-8", errors="replace")

    # Case-insensitive search for matches
    query_lower = query.lower()
    lines = text.split("\n")
    matches: list[str] = []
    citations: list[Citation] = []

    for line_num, line in enumerate(lines, 1):
        if query_lower in line.lower():
            # Find the match position within the line for context
            idx = line.lower().index(query_lower)
            start = max(0, idx - context_chars)
            end = min(len(line), idx + len(query) + context_chars)
            passage = line[start:end]

            matches.append(f"Line {line_num}: ...{passage}...")

            if len(citations) < 5:  # Cap citations at 5
                citations.append(
                    Citation(
                        chunk_id=f"text-search-{file_id}-L{line_num}",
                        file_id=file_id,
                        file_name=file_name,
                        source_url=web_link,
                        location=f"Line {line_num}",
                        snippet=passage[:300],
                    )
                )

            if len(matches) >= 20:
                break

    if not matches:
        return f"No matches for '{query}' found in {file_name}.", []

    result = f"Found {len(matches)} matches for '{query}' in {file_name}:\n\n"
    result += "\n\n".join(matches)
    return result, citations
